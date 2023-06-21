import os

from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.apps import apps

import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side

patentapp = apps.get_model('Patent', 'Patentapplication')


# Create your views here.
def dashboard(r):
    return render(r, 'admin/dashboard.html')


def adminpanel(r):
    patent = apps.get_model('Patent', 'Patentapplication')
    cp = apps.get_model('Copyright', 'Copyright')
    des = apps.get_model('Design', 'Design')
    return render(r, "admin/admin.html",
                  {'con': patent.objects.all(), 'cp': cp.objects.all(), 'des': des.objects.all()})


def Editapplication(r, uid):
    patent = patentapp.objects.get(uid=uid)
    nda = apps.get_model('Patent', 'NDAStatus')
    paymentstatus = apps.get_model('Patent', 'PaymentStatus')
    DocumentationStatus = apps.get_model('Patent', 'DocumentationStatus')
    NoveltyStatus = apps.get_model('Patent', 'NoveltyStatus')
    DraftingStatus = apps.get_model('Patent', 'DraftingStatus')
    ExaminationSatus = apps.get_model('Patent', 'ExaminationSatus')
    DrawingStatus = apps.get_model('Patent', 'DrawingStatus')
    FilingStatus = apps.get_model('Patent', 'FilingStatus')
    FerStatus = apps.get_model('Patent', 'FerStatus')
    HearingStatus = apps.get_model('Patent', 'HearingStatus')
    GrantsStatus = apps.get_model('Patent', 'GrantsStatus')
    Assignto = apps.get_model('Patent', 'Assignedto')
    assinto = Assignto.objects.get(uid=uid)
    users = apps.get_model('login', 'Usermodel').objects.all()
    context = {'app': patent, 'paymentstatus': paymentstatus.objects.filter(uid=uid),
               'ndastatus': nda.objects.get(uid=uid), 'assign': assinto, 'user': users}
    c = {}
    if patent.patenttype == 'full':
        c = {'GrantsStatus': GrantsStatus.objects.get(uid=uid),
             'DocumentationStatus': DocumentationStatus.objects.get(uid=uid),
             'FerStatus': FerStatus.objects.get(uid=uid), 'NoveltyStatus': NoveltyStatus.objects.get(uid=uid),
             'DraftingStatus': DraftingStatus.objects.get(uid=uid),
             'FilingStatus': FilingStatus.objects.get(uid=uid),
             'DrawingStatus': DrawingStatus.objects.get(uid=uid),
             'HearingStatus': HearingStatus.objects.get(uid=uid),
             'ExaminationStatus': ExaminationSatus.objects.get(uid=uid)}
    elif 'novelty search' in patent.patenttype:
        c = {'NoveltyStatus': NoveltyStatus.objects.get(uid=uid)}
    elif 'drafting' in patent.patenttype:
        c = {'DraftingStatus': DraftingStatus.objects.get(uid=uid)}
    elif 'drawing' in  patent.patenttype :
        c = {'DrawingStatus': DrawingStatus.objects.get(uid=uid)}
    elif 'documentation' in patent.patenttype:
        c = {'DocumentationStatus': DocumentationStatus.objects.get(uid=uid)}
    elif 'filing' in patent.patenttype :
        c = {'FilingStatus': FilingStatus.objects.get(uid=uid)}
    elif 'examination' in patent.patenttype:
        c = {'ExaminationSatus': ExaminationSatus.objects.get(uid=uid)}
    elif 'FER' in patent.patenttype :
        c = {'FerStatus': FerStatus.objects.get(uid=uid), }
    elif 'hearing' in patent.patenttype:
        c = {'HearingStatus': HearingStatus.objects.get(uid=uid)}
    co = {**context, **c}
    return render(r, 'Patent/editappliaction.html', co)


def ndsatatus(r):
    nda = apps.get_model('Patent', 'NDAStatus')
    if r.method == 'POST':
        uid = r.POST['uid']
        status = r.POST['status']
        ndas = nda.objects.get(uid=uid)
        ndas.status = True
        ndas.nda = status
        ndas.save()
        return redirect('user/login')


def filingstatus(r):
    FilingStatus = apps.get_model('Patent', 'FilingStatus')
    if r.method == 'POST':
        uid = r.POST['uid']
        sta = FilingStatus.objects.get(uid=uid)
        sta.status = True
        sta.save()
        return redirect('user/login')


def dashboard(r):
    return redirect('user/login')


def closeapplication(r):
    patent = apps.get_model('Patent', 'Patentapplication')
    if r.method == 'POST':
        uid = r.POST['uid']
        patent = patent.objects.get(uid=uid)
        patent.status = True
        patent.save()
        return redirect('user/login')


def openapplication(r):
    patent = apps.get_model('Patent', 'Patentapplication')
    if r.method == 'POST':
        uid = r.POST['uid']
        patent = patent.objects.get(uid=uid)
        patent.status = False
        patent.save()
        return redirect('user/login')


def EditapplicationCopyright(r, uid):
    cp = apps.get_model('Copyright', 'Copyright').objects.get(uid=uid)
    return render(r, 'Copyright/editapp.html', {'c': cp})


def track(r):
    patent = apps.get_model('Patent', 'Patentapplication')
    cp = apps.get_model('Copyright', 'Copyright')
    des = apps.get_model('Design', 'Design')
    return render(r, "admin/track.html",
                  {'con': patent.objects.all(), 'cp': cp.objects.all(), 'des': des.objects.all()})


def assignto(r):
    if r.method == 'POST':
        uid = r.POST['uid']
        ass = r.POST['assignto']
        p = apps.get_model('Patent', 'Assignedto').objects.get(uid=uid)
        p.assignto = ass
        p.assignstatus = True
        p.save()
    return redirect('user/login')


def generateworkbook(r):
    w = openpyxl.Workbook()
    patent = patentapp.objects.all()
    sheet = w.active
    sheet.merge_cells('A1:J1')
    sheet.cell(row=1, column=1).value = 'NPS completed projects Report'
    sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    row, column = 2, 1
    titles = ["Date", 'Mode of Contact', 'Reference By', 'Organization', 'Phone', 'Email',
              'Invention Title', 'Services', 'NDA', 'Assigned to Whom']
    col = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
    for i in titles:
        sheet.cell(row=2, column=column).value = i
        sheet.column_dimensions[col[column - 1]].width = 20
        sheet.column_dimensions['G'].width = 50
        sheet[col[column - 1] + str(row)].font = Font(bold=True)
        column += 1
    column = 1
    row = 3
    Assignto = apps.get_model('Patent', 'Assignedto')
    nda = apps.get_model('Patent', 'NDAStatus')
    for i in patent:
        assinto = Assignto.objects.get(uid=i.uid)
        ndastatus = nda.objects.get(uid=i.uid)
        if ndastatus.status:
            ndas = 'completed'
        else:
            ndas = ' Not Completed'
        sheet.cell(row=row, column=1).value = i.date
        sheet.cell(row=row, column=2).value = i.modeofcontact
        sheet.cell(row=row, column=3).value = i.referedby
        sheet.cell(row=row, column=4).value = i.organisation
        sheet.cell(row=row, column=5).value = i.conntactnumber
        sheet.cell(row=row, column=6).value = i.emailid
        sheet.cell(row=row, column=7).value = i.title
        sheet.cell(row=row, column=8).value = i.patenttype
        sheet.cell(row=row, column=9).value = ndas
        sheet.cell(row=row, column=10).value = assinto.assignto
        row += 1
    w.save("Reports.xlsx")
    filepath = 'Reports.xlsx'
    res = HttpResponse(open(filepath, 'rb').read(), content_type='application/xlsz')
    res['Content-Disposition'] = 'attachment;filename=' + os.path.basename(filepath)
    return res


def deleteApp(r, uid):
    patent = apps.get_model('Patent', 'Patentapplication').objects.get(uid=uid)
    apps.get_model('Patent', 'GrantsStatus').objects.get(uid=uid).delete()
    apps.get_model('Patent', 'PaymentStatus').objects.get(uid=uid).delete()
    apps.get_model('Patent', 'Assignedto').objects.get(uid=uid).delete()
    apps.get_model('Patent', 'NDAStatus').objects.get(uid=uid).delete()
    if patent.patenttype == 'full':
        apps.get_model('Patent','NoveltyStatus').objects.get(uid=uid).delete()
        apps.get_model('Patent', 'DraftingStatus').objects.get(uid=uid).delete()
        apps.get_model('Patent', 'DrawingStatus').objects.get(uid=uid).delete()
        apps.get_model('Patent', 'DocumentationStatus').objects.get(uid=uid).delete()
        apps.get_model('Patent', 'FilingStatus').objects.get(uid=uid).delete()
        apps.get_model('Patent', 'ExaminationSatus').objects.get(uid=uid).delete()
        apps.get_model('Patent', 'FerStatus').objects.get(uid=uid).delete()
        apps.get_model('Patent', 'HearingStatus').objects.get(uid=uid).delete()
    elif 'novelty search' in patent.patenttype:
        apps.get_model('Patent', 'NoveltyStatus').objects.get(uid=uid).delete()
    elif 'drafting' in patent.patenttype:
        apps.get_model('Patent', 'DraftingStatus').objects.get(uid=uid).delete()
    elif 'drawing' in patent.patenttype:
        apps.get_model('Patent', 'DrawingStatus').objects.get(uid=uid).delete()
    elif 'documentation' in patent.patenttype:
        apps.get_model('Patent', 'DocumentationStatus').objects.get(uid=uid).delete()
    elif 'filing' in patent.patenttype:
        apps.get_model('Patent', 'FilingStatus').objects.get(uid=uid).delete()
    elif 'examination' in patent.patenttype:
        apps.get_model('Patent', 'ExaminationSatus').objects.get(uid=uid).delete()
    elif 'FER' in patent.patenttype:
        apps.get_model('Patent', 'FerStatus').objects.get(uid=uid).delete()
    elif 'hearing' in patent.patenttype:
        apps.get_model('Patent', 'HearingStatus').objects.get(uid=uid).delete()
    patent.delete()
    return redirect('user/login')

